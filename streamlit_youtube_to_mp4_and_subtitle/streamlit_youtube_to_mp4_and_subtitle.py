import streamlit as st
import tkinter as tk
from tkinter import filedialog
import time
import pandas as pd
import openpyxl
import os
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from openpyxl.utils import get_column_letter
from pytube import YouTube
import re
import shutil


# ====================== Helper Functions ======================

def setup_webdriver():
    # Configuration for setting up a webdriver
    chrome_options = Options()
    user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
    chrome_options.add_argument(f"user-agent={user_agent}")
    chrome_options.add_experimental_option("detach", True)
    chrome_options.add_experimental_option('useAutomationExtension', False)
    chrome_options.add_experimental_option('excludeSwitches', ['enable-automation'])
    return webdriver.Chrome(options=chrome_options)

def select_folder():
    # UI for folder selection
    root = tk.Tk()
    root.withdraw()
    folder_path = filedialog.askdirectory(master=root)
    root.destroy()
    return folder_path

def create_directory_if_not_exists(directory_path):
    """Creates a directory if it does not exist."""
    if not os.path.exists(directory_path):
        os.makedirs(directory_path)

def sanitize_filename(filename):
    """Remove invalid characters from the filename."""
    return re.sub(r'[^\w\s\.]', '_', filename)

def move_vtt_file_to_main_folder(directory, vtt_subfolder_path):
    """Move the genuine .vtt file from its subfolder to the main folder."""
    main_folder = os.path.join(directory)

    # Search for the .vtt file inside the subfolder
    genuine_vtt_path = None
    for file_name in os.listdir(vtt_subfolder_path):
        if file_name.endswith('.vtt'):
            genuine_vtt_path = os.path.join(vtt_subfolder_path, file_name)
            break

    if genuine_vtt_path:  # if .vtt file is found
        new_vtt_path = os.path.join(main_folder, os.path.basename(genuine_vtt_path))
        shutil.move(genuine_vtt_path, new_vtt_path)

        # Delete the subfolder
        shutil.rmtree(vtt_subfolder_path)
        return new_vtt_path
    else:
        # If for some reason the genuine .vtt isn't where we expect, raise an exception.
        raise FileNotFoundError(f"Genuine .vtt file not found in {vtt_subfolder_path}")

def rename_video_to_match_vtt(vtt_path):
    directory = os.path.dirname(vtt_path)
    vtt_filename_without_extension = os.path.splitext(os.path.basename(vtt_path))[0]

    # Searching for the video in the main folder
    for file in os.listdir(directory):
        if file.endswith(".mp4"):
            os.rename(os.path.join(directory, file), os.path.join(directory, f"{vtt_filename_without_extension}.mp4"))
            break

# ====================== Main Functions ======================

def get_videos_data(driver, url):
    # Function to scrape video data from YouTube using selenium
    driver.get(url)
    channel_name = driver.find_element(By.XPATH, '//*[@id="text"]').text

    # Infinite scroll handling
    WAIT_IN_SECONDS = 5
    last_height = driver.execute_script("return document.documentElement.scrollHeight")

    while True:
        driver.execute_script("window.scrollTo(0, arguments[0]);", last_height)
        time.sleep(WAIT_IN_SECONDS)
        new_height = driver.execute_script("return document.documentElement.scrollHeight")
        if new_height == last_height:
            break
        last_height = new_height

    views = driver.find_elements(By.XPATH, '//div[@id="metadata-line"]/span[1]')
    titles = driver.find_elements(By.ID, "video-title")
    links = driver.find_elements(By.ID, "video-title-link")

    return channel_name, [{'title': t.text, 'views': v.text, 'link': l.get_attribute('href')} for t, v, l in zip(titles, views, links)]

def save_to_excel(channel_name, data, directory):
    # Function to save fetched data to an Excel file
    filename = f"{channel_name}.xlsx"
    filepath = os.path.join(directory, filename)

    df = pd.DataFrame(data)
    df.to_excel(filepath, index=False, engine='openpyxl')

    workbook = openpyxl.load_workbook(filepath)
    worksheet = workbook.active

    col_widths = [600, 150, 300]
    for i, width in enumerate(col_widths, 1):
        worksheet.column_dimensions[get_column_letter(i)].width = width / 6

    workbook.save(filepath)
    return filepath

def download_videos(filename, directory):
    # Function to download videos from YouTube links in the Excel
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active

    for row in range(2, sheet.max_row + 1):
        url = sheet.cell(row=row, column=3).value
        yt = YouTube(url)
        stream = yt.streams.get_highest_resolution()
        stream.download(directory)

def convert_to_vtt(filename, directory, model, task, language):
    # Function to convert YouTube videos to VTT format
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active

    for row in range(2, sheet.max_row + 1):
        youtube_link = sheet.cell(row=row, column=3).value
        video_title = sheet.cell(row=row, column=1).value
        sanitized_title = sanitize_filename(video_title)
        vtt_file_path = os.path.join(directory, f"{sanitized_title}.vtt")

        os.system(
            f'yt_whisper "{youtube_link}" --model {model} --task {task} --language {language} --output "{vtt_file_path}"')
        time.sleep(2)

        new_vtt_path = move_vtt_file_to_main_folder(directory, vtt_file_path)
        rename_video_to_match_vtt(new_vtt_path)

    workbook.close()

# ====================== Streamlit UI ======================

def main_streamlit():
    st.title('YouTube Video Data and Transcriptions')

    channel_url = st.text_input('Enter YouTube Channel URL:', 'https://www.youtube.com/@mongdang111/videos')

    if 'directory' not in st.session_state:
        st.session_state.directory = 'C:/Users'

    if st.button("Select Folder"):
        st.session_state.directory = select_folder() or st.session_state.directory

    st.write(f"Files will be stored in: {st.session_state.directory}")

    model = st.selectbox('Choose a model:', ['base', 'small', 'medium', 'large'])
    task = st.selectbox('Choose a task:', ['transcribe', 'translate'])
    language = st.selectbox('Choose a language:', ['en', 'ko'])

    if st.button('Fetch, Download and Convert'):
        try:
            driver = setup_webdriver()
            channel_name, videos_data = get_videos_data(driver, channel_url)
            filename = save_to_excel(channel_name, videos_data, st.session_state.directory)
            st.success(f'Successfully saved data to {filename}.')
            download_videos(filename, st.session_state.directory)
            st.success('Videos downloaded successfully.')
            convert_to_vtt(filename, st.session_state.directory, model, task, language)
            st.success('Videos converted to VTT successfully.')
        except Exception as e:
            st.error(f"An error occurred: {e}")

    st.write('Note: Ensure you have the necessary dependencies and WebDriver installed for this to work.')


if __name__ == "__main__":
    main_streamlit()
